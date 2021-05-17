import csv
import os
from pathlib import Path
from openpyxl import Workbook,load_workbook

######
# count_fails(str)
# 
# Counts how many 'errors' the string contains. The string is expectedly the unit test results string for a particular
# submitted program, and can contain either 'E' 'F' or '.' as results. E and F are counted as errors and are treated similarly.

def count_fails(str):
    return str.count('E') + str.count('F')

#####
# nofails(string)
#
# Used to check if there were any unit test fails in the test string
# string: the unit test results. Contains 'E' 'F' and '.' for each unit test performed. Any 'E' and 'F' are test fails.


def nofails(string,maxlen):
    if 'E' in string or 'F' in string or len(string) != maxlen:
        return False
    else:
        return True

#####
# grouper(dir,round,roundnumber)
#
# dir           : The directory path of the main data
# round         : The name of the round to be processed
# roundnumber   : Round # for writing the excel sheet
# 
# Creates various lists of submissions based on how many errors. Also saves an Excel file on submission numbers
#
# total_subs                : A list of All submissions
# first_subs                : A list of first submissions by each unique user
# perfect_subs              : List of flawless submissions
# perfect_first_subs_subs   : List of first submissions that were flawless right away
# perf_uniq_subs            : List of perfect submissions, but with maximum of one per user (for the amount of users who got at least one perfect)
# 
#
# One row in the CSV file is a list with the following fields: Time,UserID,Status,Grade,filename,Unittest
# First row is the field definition and does not contain actual submission data.
#

def grouper(dir,round,roundnumber):
    with open(dir + '/' + round + '.csv', 'r') as f: 
        reader = csv.reader(f, delimiter =',')
        lines = list(reader)  # Store the data in the .csv file related to the current round containing all submissions and unit test results
        
    line_count, maxlen = 0,0
    total_subs,first_subs,perfect_subs,imperfect_subs,perfect_first_subs,temp_first,temp_perf,perf_uniq_subs = [], [], [], [], [], [], [], []
    
    # The way the unit test data is displayed in the CSV-file (tests that failed to execute are blank spaces)
    # means the unit tests string has variable length, we have to scan through the whole list to find
    # the maximum length of the unit tests field (the amount of unit tests in existence for this round). This value is needed
    # in order to divide the test results into groups for further analysis.
    # This is STILL imperfect because a given round might have a set of submissions where all of them failed to execute on one or more unit tests
    #
    # If that happens, it is a super anomaly, and we'll just ignore the chance of that happening. Also buy a lottery ticket. This is more
    # of a danger with small datasets of submitted files and it doesn't make a lot of sense to run statistical analysis on small datasets anyway --Timo
    for row in lines:
        if line_count == 0: # Skip the field definition line of the CSV
            pass
        else:
            maxlen = max(maxlen,len(row[5]))
        line_count += 1
        
    line_count = 0

    round_sort_by_fails = [[]for i in range(maxlen+1)]  # List of lists of submissions sorted by the amount of unit test fails    
    for row in lines:
        if line_count == 0: # Skip the field definition line of the CSV
            pass
        else:
            total_subs.append(row) # list of all submissions
            if row[1] not in temp_first:
                first_subs.append(row) # List of first submissions by each unique user. The file is chronologically ordered
                if nofails(row[5],maxlen): # When the first submission by a user had no errors (was a perfect first-time submission)
                    perfect_first_subs.append(row)
                temp_first.append(row[1]) # Count only the first sub from each user (row[1])
            if nofails(row[5],maxlen):
                temp_perf.append(row[1])
                perfect_subs.append(row[1]) # List of submissions with no unit test fails. May contain multiples per user
            if row[5]: # A submission exists, for an empty unit test string no file is provided at all
                fails = count_fails(row[5]) # How many unit test fails that submission had
                round_sort_by_fails[fails].append(row[4]) # Add to the list of filenames with 'errors' amount of unit test fails
        line_count += 1        
    
    perf_uniq_subs = set(perfect_subs) # Perfect subs but only one per unique user.
    imperfect_subs = len(first_subs) - len(perf_uniq_subs) # Amount of users who never submitted a perfect

# Printing out the excel sheet
    
    xls_filename = dir + '-analysis/Subdata.xlsx'
    if not os.path.exists(xls_filename):  #File doesn't yet exist, create a new file
        wb = Workbook()
        sheet = wb.active
        
        # Setup the information fields and the headers on the first go
        sheet.column_dimensions['A'].width = 32
        sheet.column_dimensions['D'].width = 20
        sheet['A1'] = 'Overall Submission data'
        sheet['A2'] = 'Metric'
        sheet['B2'] = 'Amount'
        sheet['C2'] = 'Percent'
        sheet['D2'] = '% Explanation'
            
        sheet['A3'] = 'Explanation of the fields in this document'
        sheet['A4'] = 'Round'
        sheet['B4'] = 'The name of the round of exercises'

        sheet['A5'] = 'Unit Tests'
        sheet['B5'] = 'How many unit tests were run on that round of exercises'

        sheet['A6'] = 'Total'
        sheet['B6'] = 'Total amount of submissions for that round'

        sheet['A7'] = 'User First' 
        sheet['B7'] = 'The amount of unique users who submitted at least once'

        sheet['A8'] = 'Perfect First'
        sheet['B8'] = 'Amount of users who submitted a version with no unit test fails on their first submit'

        sheet['A9'] = 'Perfect'
        sheet['B9'] = 'Amount of submissions with no errors. May contain multiples by the same user'

        sheet['A10'] = 'Perfect Unique'
        sheet['B10'] = 'The amount of users who submitted a solution with no errors at least once'

        sheet['A11'] = 'Imperfect'
        sheet['B11'] = 'The amount of users who did not submit a solution with no errors'

        sheet.freeze_panes = 'A3' # Freeze the title row for viewing convenience
    else:
        wb = load_workbook(xls_filename)
        sheet = wb.active    
    

    sheet.insert_rows(idx=3, amount=8) 
    sheet['A3'] = 'Round: ' + round
    sheet['B3'] = 'Unit Tests: ' + str(maxlen)

    sheet['A4'] = 'Total Submissions'
    sheet['B4'] = len(total_subs)

    sheet['A5'] = 'User First Submissions'
    sheet['B5'] = len(first_subs)

    sheet['A6'] = 'Perfect First Submissions'
    sheet['B6'] = len(perfect_first_subs)
    sheet['C6'] = len(perfect_first_subs)/len(first_subs) *100
    sheet['D6'] = '% of unique users'
    
    sheet['A7'] = 'Total Perfect Submissions'
    sheet['B7'] = len(perfect_subs)
    sheet['C7'] = len(perfect_subs)/len(total_subs) *100
    sheet['D7'] = '% of total submissions'

    sheet['A8'] = 'Perfect unique submissions'
    sheet['B8'] = len(perf_uniq_subs)
    sheet['C8'] = len(perf_uniq_subs)/len(first_subs) *100
    sheet['D8'] = '% of unique users'
    
    sheet['A9'] = 'Imperfect Final submissions'
    sheet['B9'] = imperfect_subs
    sheet['C9'] = imperfect_subs/len(first_subs) *100
    sheet['D9'] = '% of unique users'

# Old stdout print, for testing
#    print ('Round: ' + round + '  Unit Tests: ' + str(maxlen))
#    print (' -----')
#    print ('Total submissions            : ' + str(len(total_subs)))
#    print ('User First submissions       : ' + str(len(first_subs)))
#    print ('Perfect First submissions    : ' + str(len(perfect_first_subs)) + '\t' + str(len(perfect_first_subs)/len(first_subs) *100) + '% of unique users')
#    print ('Perfect submissions          : ' + str(len(perfect_subs)) + '\t' + str(len(perfect_subs)/len(total_subs) *100) + '% of total amount of submissions')
#    print ('Perfect unique submissions   : ' + str(len(perf_uniq_subs)) + '\t' + str(len(perf_uniq_subs)/len(first_subs) *100) + '% of unique users')
#    print ('Imperfect Final submissions  : ' + str(imperfect_subs) + '\t' + str((imperfect_subs)/len(first_subs) *100) + '% of unique users')
#    print (' --------')
    
    wb.save(xls_filename)
    return(round_sort_by_fails)

#####
# sort_to_files(dir,lists,round)
#
# dir  : Base directory of the analysis data
# lists: A list of lists with filenames grouped by the amount of unit test fails
# round: name of the round the files belong to
#
# Prints the lists into files to be fed into the metrics calc, or to be used in further analysis.
# One file is printed for each round for each amount of unit test fails, listing all the submissions in that category of fails
# 

def sort_to_files(dir,lists,round):
    line_count = 0
    
    for list in lists:
        filename = dir + '-analysis/metrics/' + round + '/filelists/' + round + '-fails-' + str(line_count) + '.txt'
        if list: #not empty
            with open(filename, 'w') as f:
                for file in list:
                    f.write(file)
                    f.write('\n')
        line_count += 1

#####
# create_dirs(dir,round_list)
#
# Creates the directory structure for the analysis results
# ./*name*-analysis/plots/*round*
# ./*name*-analysis/metrics/*round*
# 
def create_dirs(dir,round_list):
    for round in round_list:
        Path(dir + '-analysis/plots/' + round).mkdir(parents=True, exist_ok=True)
        Path(dir + '-analysis/metrics/' + round).mkdir(parents=True, exist_ok=True)
        Path(dir + '-analysis/metrics/' + round + '/filelists/').mkdir(parents=True, exist_ok=True)
        Path(dir + '-analysis/metrics/' + round + '/results/').mkdir(parents=True, exist_ok=True)

######
# main()
# Only used for testing during development. All this stuff is done in analyzer.main()
#

def main():
    dir = 'Datadump'
    directory_contents = os.listdir(dir)
    round_list = []
    for item in directory_contents: # List all the subdirectories. Each one contains one round of submissions
        if os.path.isdir(dir + '/' + item):
            round_list.append(item)

    create_dirs(dir,round_list) # Build the directory structure for printing the results

    for i,round in enumerate(round_list): # Run the sorter for every round of exercises
        sorted_rounds = grouper(dir,round,i)
        sort_to_files(dir,sorted_rounds,round)


if __name__ == "__main__":
    main()

